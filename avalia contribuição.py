import re
import os
import PyPDF2
from openpyxl import Workbook

from transformers import pipeline

classifier = pipeline("zero-shot-classification", model="facebook/bart-large-mnli")

def classificar_com_transformer(texto_contribuicao, lista_temas):
    """
    Usa modelo de linguagem para classificar contribuição entre os temas.
    """
    resultado = classifier(texto_contribuicao[:1000], lista_temas)  # Limita o texto para performance
    tema_principal = resultado['labels'][0]
    score = resultado['scores'][0]
    return tema_principal, score


def extrair_temas_decreto(caminho_arquivo_decreto):
    """
    Extrai os temas do Decreto 7217 a partir de um arquivo de texto.

    Args:
        caminho_arquivo_decreto (str): O caminho para o arquivo de texto do decreto.

    Returns:
        dict: Um dicionário onde as chaves são os títulos/capítulos do decreto
              e os valores são listas de palavras-chave relevantes.
    """
    temas_decreto = {}
    try:
        with open(caminho_arquivo_decreto, 'r', encoding='utf-8') as arquivo_decreto:
            texto_decreto = arquivo_decreto.read()

        # Extrair TÍTULOS
        titulos = re.findall(r"TÍTULO\s+([IVXLCDM]+)\s+(.*?)\s*CAPÍTULO", texto_decreto, re.IGNORECASE | re.DOTALL)
        for titulo_num, titulo_texto in titulos:
            temas_decreto[f"TÍTULO {titulo_num}"] = [word.strip() for word in re.split(r'\s+|[,;.]', titulo_texto) if len(word) > 3]

        # Extrair CAPÍTULOS
        capitulos = re.findall(r"CAPÍTULO\s+([IVXLCDM]+)\s+(.*?)\s*Art\.", texto_decreto, re.IGNORECASE | re.DOTALL)
        for capitulo_num, capitulo_texto in capitulos:
            temas_decreto[f"CAPÍTULO {capitulo_num}"] = [word.strip() for word in re.split(r'\s+|[,;.]', capitulo_texto) if len(word) > 3]

    except FileNotFoundError:
        print(f"Erro: Arquivo do decreto não encontrado.")
        return {}
    except Exception as e:
        print(f"Ocorreu um erro ao processar o arquivo do decreto: {e}")
        return {}

    return temas_decreto

def ler_pdf(caminho_pdf):
    """
    Lê o texto de um arquivo PDF.

    Args:
        caminho_pdf (str): O caminho para o arquivo PDF.

    Returns:
        str: O texto completo do PDF, ou None em caso de erro.
    """
    texto = ""
    try:
        with open(caminho_pdf, 'rb') as arquivo_pdf:
            leitor_pdf = PyPDF2.PdfReader(arquivo_pdf)
            for pagina in leitor_pdf.pages:
                texto += pagina.extract_text() or ""  # Handle None returns
        return texto
    except FileNotFoundError:
        print(f"Erro: Arquivo PDF não encontrado em '{caminho_pdf}'")
        return None
    except Exception as e:
        print(f"Ocorreu um erro ao ler o PDF '{caminho_pdf}': {e}")
        return None

def avaliar_contribuicao(contribuicao, temas_decreto):
    """
    Avalia uma contribuição com base nos temas extraídos do Decreto 7217.

    Args:
        contribuicao (str): O texto da contribuição.
        temas_decreto (dict): Um dicionário de temas do decreto e palavras-chave associadas.

    Returns:
        dict: Uma avaliação da contribuição, incluindo temas relevantes and structured output.
    """
    
    avaliacao = {
        "summary": "",  # Placeholder for overall assessment
        "thematic_analysis": {},  # Detailed themes and relevance
        "legal_references": [],  # Articles, etc.
        "contribution_type": "desconhecido",
        "argument_assessment": {}  # Strength, etc.
    }
    texto_lower = contribuicao.lower()

    # Identificar o tipo de contribuição (exemplo)
    if "incoerência" in texto_lower or "atrito" in texto_lower:
        avaliacao["contribution_type"] = "identificação de incoerência/atrito"
        avaliacao["summary"] = "The contribution identifies a potential inconsistency or conflict."  # Example
    elif "sugestão" in texto_lower or "alteração" in texto_lower:
        avaliacao["contribution_type"] = "sugestão de alteração"
        avaliacao["summary"] = "The contribution proposes an alteration to the text."  # Example
    # ... other classifications ...

    # Extrair temas relevantes
    for tema, palavras_chave in temas_decreto.items():
        relevance_score = 0
        for palavra in palavras_chave:
            if palavra in texto_lower:
                relevance_score += 1
        if relevance_score > 0:
            avaliacao["thematic_analysis"][tema] = relevance_score

    # Analisar menções a artigos (exemplo)
    artigos_mencionados = re.findall(r"Art\.?\s*(\d+)", contribuicao)
    if artigos_mencionados:
        avaliacao["legal_references"] = artigos_mencionados

    # Avaliação da argumentação (exemplo - needs refinement)
    if avaliacao["contribution_type"] == "identificação de incoerência/atrito":
        if "enfraquecimento" in texto_lower or "conflito" in texto_lower:
            avaliacao["argument_assessment"]["strength"] = "moderate"
            avaliacao["argument_assessment"]["notes"] = "The argument highlights a potential weakening of the original intent."
        else:
            avaliacao["argument_assessment"]["strength"] = "weak"
    tema_semantico, confianca = classificar_com_transformer(texto_contribuicao, list(temas_decreto_7217.keys()))
    avaliacao["classificacao_semantica"] = tema_semantico
    avaliacao["confianca_semantica"] = confianca
    return avaliacao

def print_avaliacao(avaliacao):
    """
    Prints the evaluation of a contribution to the console in a formatted way.

    Args:
        avaliacao (dict): The evaluation dictionary.
    """
    print("-" * 40)
    print("Avaliação da Contribuição:")
    print("-" * 40)

    print(f"  Resumo: {avaliacao['summary']}")
    print("\n  Análise Temática:")
    if avaliacao["thematic_analysis"]:
        for tema, relevancia in avaliacao["thematic_analysis"].items():
            print(f"    - {tema}: Relevância = {relevancia}")
    else:
        print("    Nenhum tema relevante encontrado.")

    print("\n  Referências Legais:")
    if avaliacao["legal_references"]:
        print(f"    Artigos: {', '.join(avaliacao['legal_references'])}")
    else:
        print("    Nenhuma referência legal encontrada.")

    print(f"\n  Tipo de Contribuição: {avaliacao['contribution_type']}")

    print("\n  Avaliação do Argumento:")
    if avaliacao["argument_assessment"]:
        for key, value in avaliacao["argument_assessment"].items():
            print(f"    - {key}: {value}")
    else:
        print("    Nenhuma avaliação do argumento disponível.")
    print("-" * 40)

def salvar_resultados_em_excel(resultados, caminho_excel):
    """
    Salva os resultados da avaliação das contribuições em um arquivo XLSX.

    Args:
        resultados (list): Uma lista de dicionários, onde cada dicionário contém
                          o texto da contribuição e sua avaliação.
        caminho_excel (str): O caminho para o arquivo XLSX onde os resultados serão salvos.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Avaliação das Contribuições"

    # Cabeçalhos
    sheet['A1'] = "Arquivo PDF"
    sheet['B1'] = "Resumo"
    sheet['C1'] = "Tipo de Contribuição"
    sheet['D1'] = "Temas Relevantes"
    sheet['E1'] = "Relevância (por tema)"
    sheet['F1'] = "Artigos Mencionados"
    sheet['G1'] = "Força do Argumento"
    sheet['H1'] = "Notas do Argumento"
    sheet['I1'] = "Tema Semântico (Transformer)"
    sheet['J1'] = "Confiança (%)"

    sheet.cell(row=linha, column=9, value=resultado["avaliacao"].get("classificacao_semantica", ""))
    sheet.cell(row=linha, column=10, value=round(resultado["avaliacao"].get("confianca_semantica", 0) * 100, 2))


    linha = 2
    for resultado in resultados:
        sheet.cell(row=linha, column=1, value=resultado["arquivo"])
        sheet.cell(row=linha, column=2, value=resultado["avaliacao"]["summary"])
        sheet.cell(row=linha, column=3, value=resultado["avaliacao"]["contribution_type"])
        sheet.cell(row=linha, column=4, value=", ".join(resultado["avaliacao"].get("thematic_analysis", {}).keys()))
        sheet.cell(row=linha, column=5, value=str(resultado["avaliacao"].get("thematic_analysis", {})))
        sheet.cell(row=linha, column=6, value=", ".join(resultado["avaliacao"].get("legal_references", [])))
        sheet.cell(row=linha, column=7, value=resultado["avaliacao"].get("argument_assessment", {}).get("strength", ""))
        sheet.cell(row=linha, column=8, value=resultado["avaliacao"].get("argument_assessment", {}).get("notes", "") )
        linha += 1

    try:
        workbook.save(caminho_excel)
        print(f"Resultados salvos em '{caminho_excel}'")
    except Exception as e:
        print(f"Ocorreu um erro ao salvar o arquivo Excel: {e}")

if __name__ == "__main__":
    caminho_arquivo_decreto = r"W:\MINISTÉRIO DAS CIDADES\Consulta 7217\apresentação\base decreto.txt"
    pasta_contribuicoes = r"W:\MINISTÉRIO DAS CIDADES\Consulta 7217\Contribuições PDF"
    caminho_arquivo_excel = r"W:\MINISTÉRIO DAS CIDADES\Consulta 7217\avaliacao_contribuicoes.xlsx"

    temas_decreto_7217 = extrair_temas_decreto(caminho_arquivo_decreto)

    if not temas_decreto_7217:
        print("Não foi possível extrair os temas do decreto. Encerrando.")
        exit()

    resultados_avaliacao = []
    arquivos_pdf = [arquivo for arquivo in os.listdir(pasta_contribuicoes) if arquivo.lower().endswith(".pdf")]

    for arquivo_pdf in arquivos_pdf:
        caminho_completo_pdf = os.path.join(pasta_contribuicoes, arquivo_pdf)
        texto_contribuicao = ler_pdf(caminho_completo_pdf)

        if texto_contribuicao:
            avaliacao = avaliar_contribuicao(texto_contribuicao, temas_decreto_7217)
            resultados_avaliacao.append({
                "arquivo": arquivo_pdf,
                "contribuicao": texto_contribuicao,
                "avaliacao": avaliacao
            })
            print_avaliacao(avaliacao)  # Print to console
        else:
            print(f"Não foi possível ler o arquivo: {arquivo_pdf}")

    if resultados_avaliacao:
        salvar_resultados_em_excel(resultados_avaliacao, caminho_arquivo_excel)
    else:
        print("Nenhuma contribuição válida encontrada para avaliar.")