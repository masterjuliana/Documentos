import re
import json
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def parse_decreto_para_dados_tabela(texto_decreto):
    """
    Analisa o texto de um decreto para extrair informações estruturadas
    e a quantidade de contribuições por item.
    """
    dados_processados = []

    current_titulo = ""
    current_capitulo = ""
    current_secao = ""
    current_artigo = ""
    current_inciso = "" # Para manter o contexto do último Inciso

    # Dividir o texto em linhas para processamento
    linhas = texto_decreto.split('\n')

    # Expressões regulares para identificar os elementos (MAIS FLEXÍVEIS)
    # Garante que TÍTULO, CAPÍTULO e Seção sejam lidos, mesmo que tenham texto adicional.
    re_titulo = re.compile(r'^(TÍTULO [IVXLCDM]+(?:[A-ZÇÃÉÊÔÕÚÁÀÍÓÚÂÊÎÔÛÜÇÁÉÍÓÚ\s]+)?)$', re.IGNORECASE)
    re_capitulo = re.compile(r'^(CAPÍTULO [IVXLCDM]+(?:[A-ZÇÃÉÊÔÕÚÁÀÍÓÚÂÊÎÔÛÜÇÁÉÍÓÚ\s]+)?)$', re.IGNORECASE)
    re_secao = re.compile(r'^(Seção [IVXLCDM]+(?:[A-ZÇÃÉÊÔÕÚÁÀÍÓÚÂÊÎÔÛÜÇÁÉÍÓÚ\s]+)?|Seção [A-Za-zÀ-ÖØ-öø-ÿ\s]+)$', re.IGNORECASE)

    # Novo e melhor regex para capturar o item.
    # Ele busca o padrão de número_item espaço quantidade_contribuicao e então o texto.
    # O .*? é crucial para pegar o mínimo de texto até o próximo padrão ou fim da linha.
    # Adicionei `\s*` em vários lugares para flexibilidade de espaços.
    # re.DOTALL permite que `.` case também com novas linhas, mas aqui estamos processando linha por linha.
    # Por isso, o foco é em garantir que a linha de entrada esteja bem formatada.
    re_item_e_contribuicao = re.compile(
        r'^(\d+)\s*(\d+)\s*(Art\.\s*\d+[º]?|Parágrafo único\.?|\u00a7\s*\d+\u00ba?|[IVXLCDM]+[.\s]*-|[a-z]\)[\s\S]*?)(.*)$', re.IGNORECASE
    )
    # Grupo 1: Número do Item
    # Grupo 2: Quantidade de Contribuições
    # Grupo 3: O "tipo" do item (Art., §, Inciso, Alínea) - isso será o início do "Texto do Item"
    # Grupo 4: O restante do texto da linha

    # Expressões para identificar o tipo de item APÓS a captura inicial do Grupo 3
    re_artigo_interno = re.compile(r'^(Art\.\s*\d+[º]?|Artigo\s*\d+)', re.IGNORECASE)
    re_paragrafo_interno = re.compile(r'^\u00a7\s*(\d+)\u00ba?|Parágrafo único\.?', re.IGNORECASE) # §
    re_inciso_interno = re.compile(r'^[IVXLCDM]+[.\s]*-', re.IGNORECASE) # Ajustado para . ou espaço
    re_alinea_interno = re.compile(r'^[a-z]\)', re.IGNORECASE)

    for i, linha in enumerate(linhas):
        linha_limpa = linha.strip()

        if not linha_limpa:
            continue

        # Verifica TÍTULO
        match_titulo = re_titulo.match(linha_limpa)
        if match_titulo:
            current_titulo = match_titulo.group(1).strip()
            current_capitulo = ""
            current_secao = ""
            current_artigo = ""
            current_inciso = ""
            continue

        # Verifica CAPÍTULO
        match_capitulo = re_capitulo.match(linha_limpa)
        if match_capitulo:
            current_capitulo = match_capitulo.group(1).strip()
            current_secao = ""
            current_artigo = ""
            current_inciso = ""
            continue

        # Verifica SEÇÃO
        match_secao = re_secao.match(linha_limpa)
        if match_secao:
            current_secao = match_secao.group(1).strip()
            current_artigo = ""
            current_inciso = ""
            continue

        # Verifica se é uma linha de item com número e contribuições
        match_item_contrib = re_item_e_contribuicao.match(linha_limpa)
        if match_item_contrib:
            num_item = match_item_contrib.group(1)
            qtd_contribuicoes = match_item_contrib.group(2)
            
            # Combina o texto do grupo 3 e 4 para formar o texto completo do item
            texto_completo_item = (match_item_contrib.group(3) + match_item_contrib.group(4)).strip()

            # Inicializa os campos específicos do item para a linha atual
            artigo_item = ""
            paragrafo_item = ""
            inciso_item = ""
            alinea_item = ""

            # Tenta identificar o tipo de item com base no texto_completo_item
            match_art = re_artigo_interno.match(texto_completo_item)
            if match_art:
                artigo_item = match_art.group(0).strip()
                current_artigo = artigo_item # Atualiza o contexto do Artigo
                current_inciso = "" # Reseta inciso quando um novo Artigo é encontrado
            else:
                match_paragrafo = re_paragrafo_interno.match(texto_completo_item)
                if match_paragrafo:
                    paragrafo_item = match_paragrafo.group(0).strip()
                    artigo_item = current_artigo # Mantém o Artigo atual
                    current_inciso = "" # Reseta inciso
                else:
                    match_inciso = re_inciso_interno.match(texto_completo_item)
                    if match_inciso:
                        inciso_item = match_inciso.group(0).strip()
                        artigo_item = current_artigo # Mantém o Artigo atual
                        current_inciso = inciso_item # Atualiza o contexto do Inciso
                    else:
                        match_alinea = re_alinea_interno.match(texto_completo_item)
                        if match_alinea:
                            alinea_item = match_alinea.group(0).strip()
                            artigo_item = current_artigo # Mantém o Artigo atual
                            inciso_item = current_inciso # Mantém o Inciso atual
            
            # Adiciona os dados processados
            dados_processados.append({
                "Título": current_titulo,
                "Capítulo": current_capitulo,
                "Seção": current_secao,
                "Art.": artigo_item,
                "Parágrafo": paragrafo_item,
                "Inciso": inciso_item,
                "Alínea": alinea_item,
                "Número do Item": num_item,
                "Quantidade de Contribuições": qtd_contribuicoes,
                "Texto do Item": texto_completo_item # O texto completo do item capturado
            })
            continue

        # Se a linha não casar com nenhum padrão de cabeçalho ou item numerado,
        # significa que não é um item de interesse para a tabela de contribuições.
        # Portanto, ela é ignorada para a formação da tabela.
        
    return dados_processados

def gerar_tabela_analise_e_planilha(dados_tabela, nome_arquivo_xlsx="analise_decreto.xlsx"):
    """
    Gera uma tabela formatada para análise no console e salva os dados
    em um arquivo XLSX (Excel).
    """
    cabecalho = [
        "Título", "Capítulo", "Seção", "Art.", "Parágrafo", "Inciso",
        "Alínea", "Número do Item", "Quantidade de Contribuições", "Texto do Item"
    ]

    # --- Parte 1: Gerar a Tabela no Console ---
    larguras = {col: len(col) for col in cabecalho}
    for linha in dados_tabela:
        for col in cabecalho:
            conteudo_celula = str(linha.get(col, ""))
            if len(conteudo_celula) > larguras[col]:
                larguras[col] = len(conteudo_celula)

    # Limita o comprimento da exibição no console
    larguras_console = {col: min(larguras[col], 70) for col in cabecalho} # Max 70 caracteres para console

    linha_cabecalho_console = " | ".join([f"{col:<{larguras_console[col]}}" for col in cabecalho])
    print(linha_cabecalho_console)

    linha_separadora_console = "+".join(["-" * larguras_console[col] for col in cabecalho])
    print(linha_separadora_console)

    for linha in dados_tabela:
        valores_linha = []
        for col in cabecalho:
            conteudo_celula = str(linha.get(col, ""))
            # Trunca o texto para exibição no console
            if len(conteudo_celula) > larguras_console[col]:
                conteudo_celula = conteudo_celula[:larguras_console[col]-3] + "..."
            valores_linha.append(conteudo_celula)

        linha_formatada_console = " | ".join([f"{valor:<{larguras_console[cabecalho[i]]}}"
                                              for i, valor in enumerate(valores_linha)])
        print(linha_formatada_console)

    print(f"\n--- Salvando dados na planilha '{nome_arquivo_xlsx}' ---")

    # --- Parte 2: Salvar a Planilha XLSX ---
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Análise Decreto"

        ws.append(cabecalho)

        for row_data in dados_tabela:
            row_to_write = [row_data.get(col, "") for col in cabecalho]
            ws.append(row_to_write)

        # Ajusta automaticamente a largura das colunas
        for col_idx, col_name in enumerate(cabecalho):
            max_length = 0
            for cell in ws[get_column_letter(col_idx + 1)]:
                try:
                    if cell.value is not None:
                        cell_value_str = str(cell.value)
                        current_length = min(len(cell_value_str), 100) # Limite de 100 caracteres na largura do Excel
                        if current_length > max_length:
                            max_length = current_length
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width

        wb.save(nome_arquivo_xlsx)
        print(f"Planilha '{nome_arquivo_xlsx}' gerada com sucesso!")
    except Exception as e:
        print(f"Erro ao salvar o arquivo XLSX: {e}")


# --- Bloco principal de execução ---
if __name__ == "__main__":
    caminho_arquivo_decreto_json = os.path.join(os.path.dirname(__file__), 'decreto.json')
    conteudo_decreto_original = ""

    # Carrega o conteúdo do decreto do arquivo JSON (com o tratamento de erro robusto)
    try:
        with open(caminho_arquivo_decreto_json, 'r', encoding='utf-8') as f:
            data = json.load(f)
            conteudo_decreto_original = data.get('conteudo_decreto', '')
            if not conteudo_decreto_original:
                print(f"Erro: A chave 'conteudo_decreto' não foi encontrada ou está vazia no arquivo '{caminho_arquivo_decreto_json}'.")
                exit()
    except UnicodeDecodeError as ude:
        print(f"Erro de decodificação UTF-8: {ude}. Tentando ler com 'latin-1'...")
        try:
            with open(caminho_arquivo_decreto_json, 'r', encoding='latin-1') as f:
                data = json.load(f)
                conteudo_decreto_original = data.get('conteudo_decreto', '')
                if not conteudo_decreto_original:
                    print(f"Erro: A chave 'conteudo_decreto' não foi encontrada ou está vazia no arquivo '{caminho_arquivo_decreto_json}'.")
                    exit()
        except Exception as e_latin1:
            print(f"Erro fatal ao carregar o arquivo JSON com latin-1: {e_latin1}")
            exit()
    except FileNotFoundError:
        print(f"Erro: O arquivo '{caminho_arquivo_decreto_json}' não foi encontrado.")
        exit()
    except json.JSONDecodeError as jde:
        print(f"Erro: O arquivo '{caminho_arquivo_decreto_json}' não é um JSON válido: {jde}.")
        exit()
    except Exception as e:
        print(f"Ocorreu um erro inesperado ao carregar o arquivo JSON: {e}")
        exit()

    # --- REALIZAÇÃO DA LIMPEZA DO TEXTO COMPLETA ANTES DE PASSAR PARA O PARSER ---
    # ESSA É A PARTE CRÍTICA PARA GARANTIR QUE O PARSER RECEBA O TEXTO NO FORMATO CORRETO.
    conteudo_decreto_limpo = conteudo_decreto_original

    if conteudo_decreto_limpo:
        print("\n--- Iniciando processo de limpeza do texto do decreto para parsing ---")
        
        # 1. Substitui caracteres indesejados por espaço
        conteudo_decreto_limpo = conteudo_decreto_limpo.replace('\u00a0', ' ')
        conteudo_decreto_limpo = conteudo_decreto_limpo.replace('\t', ' ')
        conteudo_decreto_limpo = conteudo_decreto_limpo.replace('\r', '') # Remove retornos de carro

        # 2. Insere quebras de linha estratégicas antes de elementos estruturais
        # Isso é crucial porque o JSON pode ter achatado as linhas.
        # Adiciona quebra de linha antes de TÍTULO, CAPÍTULO, Seção, Art.
        conteudo_decreto_limpo = re.sub(r'(TÍTULO [IVXLCDM]+)', r'\n\1\n', conteudo_decreto_limpo, flags=re.IGNORECASE)
        conteudo_decreto_limpo = re.sub(r'(CAPÍTULO [IVXLCDM]+)', r'\n\1\n', conteudo_decreto_limpo, flags=re.IGNORECASE)
        conteudo_decreto_limpo = re.sub(r'(Seção [IVXLCDM]+|Seção [A-Za-zÀ-ÖØ-öø-ÿ\s]+)', r'\n\1\n', conteudo_decreto_limpo, flags=re.IGNORECASE)
        
        # Adiciona quebra de linha antes de padrões de item "NÚMERO NÚMERO Art. " ou "NÚMERO NÚMERO I - " etc.
        # Isso garante que cada item a ser parseado esteja em uma linha separada
        conteudo_decreto_limpo = re.sub(
            r'(\d+)\s*(\d+)\s*(Art\.\s*\d+[º]?|Parágrafo único\.?|\u00a7\s*\d+\u00ba?|[IVXLCDM]+[.\s]*-|[a-z]\))', 
            r'\n\1 \2 \3', # Inserir quebra de linha e manter os grupos capturados com um espaço entre eles
            conteudo_decreto_limpo, flags=re.IGNORECASE
        )


        # 3. Normaliza múltiplos espaços e quebras de linha
        conteudo_decreto_limpo = re.sub(r' +', ' ', conteudo_decreto_limpo) # Múltiplos espaços para um único
        conteudo_decreto_limpo = re.sub(r'\n{2,}', '\n', conteudo_decreto_limpo) # Múltiplas quebras de linha para uma única
        
        # 4. Remove o cabeçalho e rodapé do site que não são parte do decreto
        # Use o texto original do PDF ou site para ajustar essas regexes, se necessário.
        # Estes são exemplos baseados no texto que já vi antes.
        conteudo_decreto_limpo = re.sub(r'Presidência da República.*?Página Inicial \/ Órgãos Públicos.*?regulamentador da Lei nº 11\.445, de 2007\.', '', conteudo_decreto_limpo, flags=re.DOTALL)
        conteudo_decreto_limpo = re.sub(r'Órgão: Ministério das Cidades.*?Desde já agradecemos a sua participação!\nConteúdo\n- Clique no balão.*?contribuição -', '', conteudo_decreto_limpo, flags=re.DOTALL)
        
        # Remove quaisquer espaços em branco no início e no final da string completa.
        conteudo_decreto_limpo = conteudo_decreto_limpo.strip()

        print("--- Limpeza do texto para parsing concluída. ---")
        print(f"\nTexto limpo para parsing (primeiros 1000 caracteres):\n{conteudo_decreto_limpo[:1000]}...")
        print(f"\nTamanho total do texto do decreto limpo: {len(conteudo_decreto_limpo)} caracteres.")

    # Analisa o decreto com o texto limpo
    dados_analisados = parse_decreto_para_dados_tabela(conteudo_decreto_limpo)

    if not dados_analisados:
        print("\nNenhum item com número e contribuições foi encontrado após o parsing. Verifique as regexes ou o formato do texto.")
    else:
        # Gera a tabela no console e salva no arquivo Excel
        gerar_tabela_analise_e_planilha(dados_analisados, nome_arquivo_xlsx="analise_decreto_gerada.xlsx")
        print("\n--- Fim do Processamento ---")