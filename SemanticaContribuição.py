import re
import os
import PyPDF2
from openpyxl import Workbook
from transformers import pipeline

# Carregar classificador semântico (zero-shot)
classifier = pipeline("zero-shot-classification", model="facebook/bart-large-mnli")

def extrair_temas_decreto(caminho_arquivo_decreto):
    temas_decreto = {}
    try:
        with open(caminho_arquivo_decreto, 'r', encoding='utf-8') as arquivo:
            texto = arquivo.read()

        titulos = re.findall(r"TÍTULO\s+([IVXLCDM]+)\s+(.*?)\s*CAPÍTULO", texto, re.IGNORECASE | re.DOTALL)
        for num, texto in titulos:
            temas_decreto[f"TÍTULO {num}"] = [w.strip() for w in re.split(r'\s+|[,;.:\-]', texto) if len(w) > 3]

        capitulos = re.findall(r"CAPÍTULO\s+([IVXLCDM]+)\s+(.*?)\s*Art\.", texto, re.IGNORECASE | re.DOTALL)
        for num, texto in capitulos:
            temas_decreto[f"CAPÍTULO {num}"] = [w.strip() for w in re.split(r'\s+|[,;.:\-]', texto) if len(w) > 3]
    except Exception as e:
        print(f"Erro ao processar decreto: {e}")
    return temas_decreto

def ler_pdf(caminho_pdf):
    texto = ""
    try:
        with open(caminho_pdf, 'rb') as f:
            leitor = PyPDF2.PdfReader(f)
            for pagina in leitor.pages:
                texto += pagina.extract_text() or ""
        return texto
    except Exception as e:
        print(f"Erro ao ler PDF {caminho_pdf}: {e}")
        return None

def classificar_com_transformer(texto, lista_temas):
    try:
        resultado = classifier(texto[:1000], lista_temas)
        return resultado['labels'][0], resultado['scores'][0]
    except Exception as e:
        print(f"Erro no classificador transformer: {e}")
        return "erro", 0.0

def avaliar_contribuicao(texto, temas_decreto):
    resultado = {
        "summary": "",
        "contribution_type": "desconhecido",
        "thematic_analysis": {},
        "legal_references": [],
        "argument_assessment": {},
    }

    texto_lower = texto.lower()

    if "incoerência" in texto_lower or "atrito" in texto_lower:
        resultado["contribution_type"] = "incoerência/atrito"
        resultado["summary"] = "Aponta possível conflito ou incoerência."
    elif "sugestão" in texto_lower or "alteração" in texto_lower:
        resultado["contribution_type"] = "sugestão de alteração"
        resultado["summary"] = "Propõe mudanças no texto."
    else:
        resultado["summary"] = "Classificação não definida com base em palavras-chave."

    for tema, palavras in temas_decreto.items():
        score = sum(1 for p in palavras if p.lower() in texto_lower)
        if score > 0:
            resultado["thematic_analysis"][tema] = score

    artigos = re.findall(r"Art\.?\s*(\d+)", texto)
    if artigos:
        resultado["legal_references"] = artigos

    if "conflito" in texto_lower or "enfraquecimento" in texto_lower:
        resultado["argument_assessment"] = {
            "strength": "moderada",
            "notes": "Aponta impacto negativo ou conflito direto com a lei."
        }

    return resultado

def salvar_resultados_excel(resultados, caminho_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Análise Contribuições"

    headers = [
        "Arquivo PDF", "Resumo", "Tipo", "Temas Relevantes", "Relevância",
        "Artigos", "Força do Argumento", "Notas", "Tema Semântico", "Confiança (%)"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for i, r in enumerate(resultados, start=2):
        a = r["avaliacao"]
        ws.cell(row=i, column=1, value=r["arquivo"])
        ws.cell(row=i, column=2, value=a.get("summary", ""))
        ws.cell(row=i, column=3, value=a.get("contribution_type", ""))
        ws.cell(row=i, column=4, value=", ".join(a.get("thematic_analysis", {}).keys()))
        ws.cell(row=i, column=5, value=str(a.get("thematic_analysis", {})))
        ws.cell(row=i, column=6, value=", ".join(a.get("legal_references", [])))
        ws.cell(row=i, column=7, value=a.get("argument_assessment", {}).get("strength", ""))
        ws.cell(row=i, column=8, value=a.get("argument_assessment", {}).get("notes", ""))
        ws.cell(row=i, column=9, value=a.get("classificacao_semantica", ""))
        ws.cell(row=i, column=10, value=round(a.get("confianca_semantica", 0) * 100, 2))

    wb.save(caminho_excel)
    print(f"\n✅ Resultados salvos em: {caminho_excel}")

# ========== EXECUÇÃO ==========

if __name__ == "__main__":
    caminho_decreto = r"W:\MINISTÉRIO DAS CIDADES\Consulta 7217\apresentação\base decreto.txt"
    pasta_pdfs = r"W:\MINISTÉRIO DAS CIDADES\Consulta 7217\Contribuições PDF"
    saida_excel = r"W:\MINISTÉRIO DAS CIDADES\Consulta 7217\avaliacao_contribuicoes_semantica.xlsx"

    temas_decreto = extrair_temas_decreto(caminho_decreto)
    lista_temas = list(temas_decreto.keys())

    resultados = []

    for nome_arquivo in os.listdir(pasta_pdfs):
        if nome_arquivo.lower().endswith(".pdf"):
            caminho_pdf = os.path.join(pasta_pdfs, nome_arquivo)
            texto = ler_pdf(caminho_pdf)
            if texto:
                aval = avaliar_contribuicao(texto, temas_decreto)
                tema_sem, conf = classificar_com_transformer(texto, lista_temas)
                aval["classificacao_semantica"] = tema_sem
                aval["confianca_semantica"] = conf
                resultados.append({
                    "arquivo": nome_arquivo,
                    "avaliacao": aval
                })

    if resultados:
        salvar_resultados_excel(resultados, saida_excel)
    else:
        print("⚠️ Nenhum PDF processado com sucesso.")
